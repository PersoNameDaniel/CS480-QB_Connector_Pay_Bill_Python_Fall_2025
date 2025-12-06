"""Excel extraction helpers for account debit worksheets."""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional
from datetime import date as DateType, datetime

from openpyxl import load_workbook  # type: ignore[import-untyped]

from .models import BillPayment


def _default_company_workbook() -> Path:
    # company_data.xlsx expected in project root (one level above src/)
    return Path(__file__).resolve().parents[1] / "company_data.xlsx"


def _normalize(h: object) -> str:
    return str(h).strip() if h is not None else ""


def _read_account_debit_sheet(
    workbook_path: Path, sheet_name: str
) -> List[BillPayment]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        try:
            ws = wb[sheet_name]
        except KeyError as exc:
            raise ValueError(f"Worksheet '{sheet_name}' not found in workbook") from exc

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return []

        headers = [_normalize(h) for h in header_row]
        index = {h.lower(): i for i, h in enumerate(headers)}

        def _get(row: tuple, *names: str):
            for name in names:
                idx = index.get(name.lower())
                if idx is not None and idx < len(row):
                    return row[idx]
            return None

        payments: List[BillPayment] = []
        for row in rows:
            # -------------------------------------
            # Read the Comments column
            # -------------------------------------
            comments = _get(row, "Comments", "Comment", "comments", "comment")
            comments_str = str(comments).strip().lower() if comments else ""

            # -------------------------------------
            # SKIP SHIPPING CHARGES rows
            # -------------------------------------
            if comments_str == "shipping charges":
                print(f"Skipping row with Shipping Charges: {row}")
                continue

            # -------------------------------------
            # Parent ID - Child ID -> take only parent (left of " - ")
            parent_child = _get(
                row, "Parent ID - Child ID", "Parent ID", "ParentID", "Parent"
            )
            if parent_child in (None, ""):
                # try alternative columns that may contain parent id
                parent_child = _get(row, "Parent ID", "Parent")

            parent_str = ""
            if parent_child not in (None, ""):
                parent_text = str(parent_child).strip()
                if " - " in parent_text:
                    parent_str = parent_text.split(" - ", 1)[0].strip()
                else:
                    parent_str = parent_text

            bank_date = _get(row, "Bank Date")
            check_amount = _get(row, "Check Amount")
            vendor_raw = _get(row, "Supplier", "Supplier Name", "Vendor", "Vendor Name")
            vendor = str(vendor_raw).strip() if vendor_raw not in (None, "") else None

            # Require amount to create a payment
            if check_amount in (None, ""):
                continue

            # convert amount
            try:
                amount_value = float(str(check_amount).strip())
            except (ValueError, TypeError):
                continue

            # -------------------------------------
            # Convert Bank Date to datetime.date
            # -------------------------------------
            parsed_date: Optional[DateType]

            if isinstance(bank_date, datetime):
                parsed_date = bank_date.date()
            elif isinstance(bank_date, DateType):
                parsed_date = bank_date
            elif isinstance(bank_date, str):
                s = bank_date.strip()
                parsed_date = None

                # Try ISO-like first 10 chars (handles "YYYY-MM-DD HH:MM:SS")
                try:
                    parsed_date = DateType.fromisoformat(s[:10])
                except ValueError:
                    parsed_date = None

                if parsed_date is None:
                    for fmt in (
                        "%Y-%m-%d",
                        "%Y-%m-%d %H:%M:%S",
                        "%m/%d/%Y",
                        "%m/%d/%y",
                    ):
                        try:
                            parsed_date = datetime.strptime(s, fmt).date()
                            break
                        except ValueError:
                            continue
            else:
                parsed_date = None

            if parsed_date is None:
                print(f"Skipping row with invalid or missing Bank Date: {bank_date!r}")
                continue

            payments.append(
                BillPayment(
                    id=parent_str,
                    date=parsed_date,  # <- guaranteed datetime.date
                    amount_to_pay=amount_value,
                    vendor=vendor,
                )
            )
        return payments
    finally:
        wb.close()


def extract_account_debit_vendor(workbook_path: Path) -> List[BillPayment]:
    """Read 'account debit vendor' and return BillPayment list using parent id and default bank."""
    return _read_account_debit_sheet(workbook_path, "account debit vendor")


def extract_account_debit_nonvendor(workbook_path: Path) -> List[BillPayment]:
    """Read 'account debit nonvendor' and return BillPayment list using parent id and default bank."""
    return _read_account_debit_sheet(workbook_path, "account debit nonvendor")


__all__ = [
    "extract_account_debit_vendor",
    "extract_account_debit_nonvendor",
]


if __name__ == "__main__":  # pragma: no cover - manual invocation
    import sys

    try:
        wb = Path("company_data.xlsx")
        vendor_rows = extract_account_debit_vendor(wb)
        nonvendor_rows = extract_account_debit_nonvendor(wb)

        print(f"Vendor rows: {len(vendor_rows)}")
        for p in vendor_rows[:10]:
            print(p)

        print(f"\nNon-vendor rows: {len(nonvendor_rows)}")
        for p in nonvendor_rows[:10]:
            print(p)
    except Exception as e:
        print(f"Error: {e}")
        print(
            "Usage: python src/excel_reader.py (run from project root where company_data.xlsx lives)"
        )
        sys.exit(1)
